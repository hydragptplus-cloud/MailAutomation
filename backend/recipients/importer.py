import csv
import re
from io import BytesIO, StringIO
from openpyxl import load_workbook
from django.db import transaction
from .models import Recipient
from common.models import Organization

MAX_IMPORT_ROWS = 10000

EMAIL_REGEX = re.compile(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}')

ALIASES = {
    "name": ["name", "full_name", "full name", "recipient", "contact name", "contact", "person name", "first name", "last name"],
    "email": ["email", "emails", "email_address", "email address", "e-mail", "mail"],
    "company": ["company", "company name", "company_name", "organization", "organisation", "org"],
    "phone": ["phone", "phones", "phone number", "phone_number", "mobile", "telephone"],
    "website": ["website", "site", "web", "url"],
    "facebook": ["facebook", "fb"],
    "instagram": ["instagram", "ig", "insta"],
    "linkedin": ["linkedin"],
    "twitter": ["twitter", "x"],
    "youtube": ["youtube", "yt"],
}

def _extract_clean_email(raw_val):
    if not raw_val:
        return ""
    match = EMAIL_REGEX.search(str(raw_val))
    return match.group(0).lower() if match else ""

def _extract_clean_phone(raw_val):
    if not raw_val:
        return ""
    cleaned = str(raw_val).strip().lstrip("'").lstrip('"').strip()
    return cleaned

def _normalize(row):
    normalized = {str(k).strip().lower(): (v or "") for k, v in row.items() if k is not None}
    result = {}
    for target, aliases in ALIASES.items():
        result[target] = next((normalized.get(a, "") for a in aliases if normalized.get(a, "") != ""), "")
    return result

@transaction.atomic()
def import_recipients(file_obj, recipient_list):
    name = file_obj.name.lower()
    rows = []
    if name.endswith(".csv"):
        text = file_obj.read().decode("utf-8-sig")
        rows = list(csv.DictReader(StringIO(text)))
    elif name.endswith((".xlsx", ".xlsm")):
        wb = load_workbook(BytesIO(file_obj.read()), read_only=True, data_only=True)
        ws = wb.active
        values = list(ws.iter_rows(values_only=True)) if ws is not None else []
        if values:
            headers = [str(x).strip() if x is not None else "" for x in values[0]]
            rows = [dict(zip(headers, row)) for row in values[1:]]
    else:
        raise ValueError("Only CSV and XLSX files are supported.")

    if len(rows) > MAX_IMPORT_ROWS:
        raise ValueError(f"Import cannot exceed {MAX_IMPORT_ROWS} rows.")

    normalized_rows = []
    incoming_emails = set()
    for row in rows:
        data = _normalize(row)
        email = _extract_clean_email(data.get("email"))
        normalized_rows.append((data, email))
        if email:
            incoming_emails.add(email)
    existing = set(
        Recipient.objects.filter(recipient_list=recipient_list, email__in=incoming_emails)
        .values_list("email", flat=True)
    )
    prospective = len(incoming_emails - existing)
    organization = Organization.objects.select_for_update().get(pk=recipient_list.organization_id)
    current_recipient_count = Recipient.objects.filter(organization=organization).count()
    if current_recipient_count + prospective > organization.max_recipients:
        raise ValueError("Recipient limit reached for this account.")

    created = updated = skipped = 0
    for data, email in normalized_rows:

        if not email:
            skipped += 1
            continue

        company_val = str(data.get("company", "")).strip()
        name_val = str(data.get("name", "")).strip() or company_val or email.split("@")[0]
        phone_val = _extract_clean_phone(data.get("phone", ""))

        # Collect social profiles & web metadata
        meta = {}
        for social_key in ["facebook", "instagram", "linkedin", "twitter", "youtube"]:
            val = str(data.get(social_key, "")).strip()
            if val:
                meta[social_key] = val

        website_val = str(data.get("website", "")).strip() or None

        # Tags
        tags = ["GoogleMapsLead"]
        if website_val:
            tags.append("HasWebsite")
        if meta.get("linkedin"):
            tags.append("HasLinkedIn")

        _, was_created = Recipient.objects.update_or_create(
            recipient_list=recipient_list,
            email=email,
            defaults={
                "organization": organization,
                "name": name_val,
                "company": company_val,
                "phone": phone_val,
                "website": website_val,
                "tags": tags,
                "metadata": meta,
            },
        )
        created += int(was_created)
        updated += int(not was_created)

    return {
        "imported_count": created + updated,
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "duplicate_count": updated,
        "invalid_count": skipped,
    }
